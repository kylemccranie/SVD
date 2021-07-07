import openpyxl
import csv
from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill
import xlsxwriter
from openpyxl.styles import Font, Color, colors, Fill 

Title = []
IssueType = []
ComponentVariant = []
DRNumber = []
IssueKey = []

numberOfRows = 0


#Styling
##def style():
##    wb = load_workbook('template_out.xlsx')
##    ws = wb.active
##   
    ##fill = PatternFill(fill_type = 'solid',
##                    ##start_color = 'f88379',
##                   ## end_color = 'f88379')
##    cell_range = ws['A1':'L1']
##    cell_range.PatternFill(fill_type = 'solid',
##                    start_color = 'f88379',
 ##                   end_color = 'f88379') 
#caluculate max number of rows in excel sheet for later functions
def calcNumberOfRows():
    global numberOfRows
    wb = load_workbook(filename = 'file.xlsx')
    ws = wb.active
    for row in range(2,ws.max_row + 1):
           numberOfRows = numberOfRows + 1
    return numberOfRows
    


#code converting csv to xlsx so that we can use openpyxl
def convert():
    wb = openpyxl.Workbook()
    ws = wb.active

    with open('JiraExport.csv') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)

    wb.save('file.xlsx')


#setting up the template sheet
def create_Template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "Type"
    ws['B1'] = "DR Number"
    ws['C1'] = "Jira Issue Key"
    ws['D1'] = "Component"
    ws['E1'] = "CVV Build #"
    ws['F1'] = "Title"
    ws['G1'] = "Description"
    ws['H1'] = "Problem Path"
    ws['I1'] = "Acceptance Criteria"
    ws['J1'] = "UI Affected (If yes then how?)"
    ws['K1'] = "Change Summary"
    ws['L1'] = "Fix Comments"
    wb.save('template_out.xlsx')
    
            
#read and store 'Summary/title' into Title Array
def readToTitle():
    wb = load_workbook(filename = 'file.xlsx')
    ws = wb.active
    for col in ws.iter_cols(min_row=2, max_col=1, max_row=numberOfRows+1, values_only = True):
        for cell in col:
            Title.append(cell)

##Writing Summary/Title Array to template sheet
def writeTitleToTemplate():
    x = 2 
    wb = load_workbook(filename = 'template_out.xlsx')
    ws = wb.active
    for i in Title:
        ws.cell(row = x, column = 6).value = i
        x = x + 1
    wb.save('template_out.xlsx')


        
#read and store 'issue type' into  Array
def readToIsssueType():
    wb = load_workbook(filename = 'file.xlsx')
    ws = wb.active
    for col in ws.iter_cols(min_row=2,min_col = 4, max_col=4, max_row=numberOfRows+1, values_only = True):
        for cell in col:
            IssueType.append(cell)

##Writing IssueType Array to template sheet
def writeIssueTypeToTemplate():
    x = 2 
    wb = load_workbook(filename = 'template_out.xlsx')
    ws = wb.active
    for i in IssueType:
        ws.cell(row = x, column = 1).value = i
        x = x + 1
    wb.save('template_out.xlsx')

##Writing Component Variant from export to Component Variant Array
def readToComponentVariant():
    wb = load_workbook(filename = 'file.xlsx')
    ws = wb.active
    for col in ws.iter_cols(min_row=2,min_col = 65, max_col=65, max_row=numberOfRows+1, values_only = True):
        for cell in col:
            ComponentVariant.append(cell)

##Writing Component Variant Array to template
def writeComponentVariantToTemplate():
    x = 2 
    wb = load_workbook(filename = 'template_out.xlsx')
    ws = wb.active
    for i in ComponentVariant:
        ws.cell(row = x, column = 4).value = i
        x = x + 1
    wb.save('template_out.xlsx')


##Writing DRNumber from export to DRNumber Array
def readToDRNumber():
    wb = load_workbook(filename = 'file.xlsx')
    ws = wb.active
    for col in ws.iter_cols(min_row=2,min_col = 71, max_col=71, max_row=numberOfRows+1, values_only = True):
        for cell in col:
            DRNumber.append(cell)

##Writing Component Variant Array to template
def writeDRNumberToTemplate():
    x = 2 
    wb = load_workbook(filename = 'template_out.xlsx')
    ws = wb.active
    for i in DRNumber:
        ws.cell(row = x, column = 2).value = i
        x = x + 1
    wb.save('template_out.xlsx')


##Writing issueKey from export to IssueKey Array
def readToIssueKey():
    wb = load_workbook(filename = 'file.xlsx')
    ws = wb.active
    for col in ws.iter_cols(min_row=2,min_col = 2, max_col=2, max_row=numberOfRows+1, values_only = True):
        for cell in col:
            IssueKey.append(cell)

##Writing Component Variant Array to template
def writeIssueKeyToTemplate():
    x = 2 
    wb = load_workbook(filename = 'template_out.xlsx')
    ws = wb.active
    for i in IssueKey:
        ws.cell(row = x, column = 3).value = i
        x = x + 1
    wb.save('template_out.xlsx')


##Test Function for inserting values
##def test():
    ##wb = load_workbook(filename = 'template_out.xlsx')
    ##ws = wb.active
    ##ws.cell(row=2, column=2).value = 2
    ##wb.save('template_out.xlsx')

        


convert()
calcNumberOfRows()
create_Template()
readToTitle()
writeTitleToTemplate()
readToIsssueType()
writeIssueTypeToTemplate()
readToComponentVariant()
writeComponentVariantToTemplate()
readToDRNumber()
writeDRNumberToTemplate()
readToIssueKey()
writeIssueKeyToTemplate()
##style()
